import os
import cv2
import numpy as np
import openpyxl
from scipy import ndimage

from common import constants
from pywt import dwt2
import pandas as pd
from skimage import metrics


def read_association():
    with open(constants.ASSOCIATION_PATH, "r") as f:
        associations = [tuple(map(str, i.rstrip().split(','))) for i in f]
    f.close()
    return associations


# remove the contents of a directory
def remove_contents(path):
    files = os.listdir(path)
    for f in files:
        os.remove(os.path.join(path, f))


# add color to gray-scale image
def add_color(gray_image):
    b, g, r = cv2.split(gray_image)
    np.multiply(b, 2, out=b, casting="unsafe")
    np.multiply(g, 0.5, out=g, casting="unsafe")
    np.multiply(r, 1.25, out=r, casting="unsafe")
    merged_channels = cv2.merge([b, g, r])
    colored_image = cv2.cvtColor(merged_channels, cv2.COLOR_BGR2RGB)
    return colored_image


# apply zer_padding on image based on dimension
def zero_padding(img, dimension_type, ref_dim):
    ht, wd = img.shape
    if dimension_type == constants.DIMENSION_HEIGHT:
        # create new image of desired size for padding
        result = np.zeros((ref_dim, wd), dtype=np.uint8)
        # compute center offset
        delta_y = (ref_dim - ht) // 2
        # copy img image into center of results image
        result[delta_y:delta_y + ht, :] = img
        # results[7:, :] = img
        # results[:ht, :] = img

    elif dimension_type == constants.DIMENSION_WIDTH:
        # create new image of desired size for padding
        result = np.zeros((ht, ref_dim), dtype=np.uint8)
        # compute center offset        
        delta_x = (ref_dim - wd) // 2
        # copy img image into center of results image
        result[:, delta_x:delta_x + wd] = img
    return result


# check if a string is "" or None
def isEmpty(str):
    if str and str.strip():
        # str is not None AND str is not empty or blank
        return False
    # str is None OR str is empty or blank
    return True


# computes dice coefficient.
def dice_score(first_image, second_image):
    volume_sum = first_image.sum() + second_image.sum()
    if volume_sum == 0:
        return np.NaN
    volume_intersect = (first_image & second_image).sum()
    return 2 * volume_intersect / volume_sum

# computes epicardium dice coefficient.
def compute_epicardium_dice_score(first_image, second_image):
    first_img_nm = np.zeros(first_image.shape)
    first_img_nm = cv2.normalize(first_image, first_img_nm, 0, 2, cv2.NORM_MINMAX)

    second_img_nm = np.zeros(second_image.shape)
    second_img_nm = cv2.normalize(second_image, second_img_nm, 0, 2, cv2.NORM_MINMAX)

    first_img_nm[first_img_nm == 2] = 1
    second_img_nm[second_img_nm == 2] = 1

    epicardium_dice_score = dice_score(first_img_nm, second_img_nm)
    return epicardium_dice_score

# computes epicardium dice coefficient.
def compute_endocardium_dice_score(first_image, second_image):
    first_img_nm = np.zeros(first_image.shape)
    first_img_nm = cv2.normalize(first_image, first_img_nm, 0, 2, cv2.NORM_MINMAX)

    second_img_nm = np.zeros(second_image.shape)
    second_img_nm = cv2.normalize(second_image, second_img_nm, 0, 2, cv2.NORM_MINMAX)

    first_img_nm[first_img_nm == 2] = 0
    second_img_nm[second_img_nm == 2] = 0

    endocardium_dice_score = dice_score(first_img_nm, second_img_nm)
    return endocardium_dice_score

# computes myocardium dice coefficient.
def compute_myocardium_dice_score(first_image, second_image):
    first_img_nm = np.zeros(first_image.shape)
    first_img_nm = cv2.normalize(first_image, first_img_nm, 0, 2, cv2.NORM_MINMAX)

    second_img_nm = np.zeros(second_image.shape)
    second_img_nm = cv2.normalize(second_image, second_img_nm, 0, 2, cv2.NORM_MINMAX)

    first_img_nm[first_img_nm == 1] = 0
    first_img_nm[first_img_nm == 2] = 1

    second_img_nm[second_img_nm == 1] = 0
    second_img_nm[second_img_nm == 2] = 1

    myocardium_dice_score = dice_score(first_img_nm, second_img_nm)
    return myocardium_dice_score

def compute_avg_dice_score(first_image, second_image):
    epi_dice_score = compute_epicardium_dice_score(first_image, second_image)
    myo_dice_score = compute_myocardium_dice_score(first_image, second_image)
    return (epi_dice_score, myo_dice_score, (epi_dice_score + myo_dice_score) / 2)

def compute_myocardium_hausdorff(first_image, second_image):
    # first_image = cv2.cvtColor(first_image, cv2.COLOR_RGB2GRAY)
    # second_image = cv2.cvtColor(second_image, cv2.COLOR_RGB2GRAY)

    first_img_nm = np.zeros(first_image.shape)
    first_img_nm = cv2.normalize(first_image, first_img_nm, 0, 2, cv2.NORM_MINMAX)

    second_img_nm = np.zeros(second_image.shape)
    second_img_nm = cv2.normalize(second_image, second_img_nm, 0, 2, cv2.NORM_MINMAX)

    first_img_nm[first_img_nm == 1] = 0
    first_img_nm[first_img_nm == 2] = 1

    second_img_nm[second_img_nm == 1] = 0
    second_img_nm[second_img_nm == 2] = 1

    # return max(directed_hausdorff(first_image, second_image)[0], directed_hausdorff(second_image, first_image)[0])
    return metrics.hausdorff_distance(first_img_nm, second_img_nm)

def compute_epicardium_hausdorff(first_image, second_image):
    # first_image = cv2.cvtColor(first_image, cv2.COLOR_RGB2GRAY)
    # second_image = cv2.cvtColor(second_image, cv2.COLOR_RGB2GRAY)

    first_img_nm = np.zeros(first_image.shape)
    first_img_nm = cv2.normalize(first_image, first_img_nm, 0, 2, cv2.NORM_MINMAX)

    second_img_nm = np.zeros(second_image.shape)
    second_img_nm = cv2.normalize(second_image, second_img_nm, 0, 2, cv2.NORM_MINMAX)

    first_img_nm[first_img_nm == 2] = 1
    second_img_nm[second_img_nm == 2] = 1

    return metrics.hausdorff_distance(first_img_nm, second_img_nm)

def compute_avg_hausdorff(first_image, second_image):
    epi_hausdorff = compute_epicardium_hausdorff(first_image, second_image)
    myo_hausdorff = compute_myocardium_hausdorff(first_image, second_image)
    return (epi_hausdorff, myo_hausdorff, (epi_hausdorff + myo_hausdorff) / 2)

def calculate_entropy(img):
    marg = np.histogramdd(np.ravel(img), bins=256)[0] / img.size
    marg = list(filter(lambda p: p > 0, np.ravel(marg)))
    entropy = -np.sum(np.multiply(marg, np.log2(marg)))
    return round(entropy, 4)

def calculate_energy(img):
    im = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    _, (cH, cV, cD) = dwt2(im.T, 'db1')
    # a - LL, h - LH, v - HL, d - HH as in matlab
    energy = (cH ** 2 + cV ** 2 + cD ** 2).sum() / im.size
    return round(energy, 4)

def calculate_mean(img):
    mean = np.mean(img)
    return round(mean, 4)

def calculate_variance(img):
    variance = ndimage.variance(img)
    return round(variance, 4)

def calculate_min_max(img):
    min = ndimage.minimum(img)
    max = ndimage.maximum(img)
    return (min, max)

def calculate_michelso_contrast(img):
    (min, max) = calculate_min_max(img)
    michelso_contrast = (max - min) // (max + min)
    return michelso_contrast


def calculate_contrast_ratio(img):
    # im = Image.fromarray(img)
    # statestic = ImageStat.Stat(im)
    # for band, name in enumerate(im.getbands()):
    # print(f'Band: {name}, min/max: {stats.extrema[band]}, stddev: {stats.stddev[band]}')
    # contrast = statestic.stddev[band]
    # print("PIL:" + str(contrast))
    contrast = img.std()
    return round(contrast, 4)


def append_df(excel_path_file, df):
    with open(excel_path_file, 'a+') as f:
        df.to_csv(f, header=f.tell() == 0, encoding='utf-8', index=False)


def get_statstics(img, file_name):
    excel_path_file = '../common/metrics/excel/final_stats.xlsx'

    entropy = calculate_entropy(img)
    energy = calculate_energy(img)
    mean = calculate_mean(img)
    variance = calculate_variance(img)
    (min, max) = calculate_min_max(img)
    michelso_contrast = calculate_michelso_contrast(img)
    contrast = calculate_contrast_ratio(img)

    df = pd.DataFrame({'File Name': [file_name],
                       'Entropy': [entropy],
                       'Energy': [energy],
                       'Mean': [mean],
                       'Variance': [variance],
                       'Min': [min],
                       'Max': [max],
                       'Michelso Contrast': [michelso_contrast],
                       'Contrast': [contrast]
                       })
    if not os.path.exists(excel_path_file):
        df.to_excel(excel_path_file, sheet_name='All', index=False)
    else:
        wb = openpyxl.load_workbook(excel_path_file)
        sheet = wb.active
        max_column = sheet.max_column
        max_row = sheet.max_row
        new_row = [file_name, entropy, energy, mean, variance, min, max, michelso_contrast, contrast]
        for j in range(max_column):
            cell = sheet.cell(max_row + 1, j + 1)
            cell.value = new_row[j]

        wb.save(excel_path_file)

def get_dice_hasudorff_distances(dice,hausdorff,file_name,sheet_name):
    excel_path_file = '../../common/metrics/excel/statestics.xlsx'
    dice_epi = round(dice[0],3)
    dice_myo = round(dice[1],3)
    dice_avg = round(dice[2],3)
    hd_epi = round(hausdorff[0],3)
    hd_myo = round(hausdorff[1],3)
    hd_avg = round(hausdorff[2],3)

    df = pd.DataFrame({'file': [file_name],
                        'dice_epi': [dice_epi],
                       'dice_myo': [dice_myo],
                       'dice_avg': [dice_avg],
                       'hdf_epi': [hd_epi],
                       'hdf_myo': [hd_myo],
                       'hdf_avg': [hd_avg]
                       })
    if not os.path.exists(excel_path_file):
        df.to_excel(excel_path_file, sheet_name=sheet_name, index=False)
    else:
        wb = openpyxl.load_workbook(excel_path_file)
        sheet = wb.active
        max_column = sheet.max_column
        max_row = sheet.max_row
        new_row = [file_name, dice_epi, dice_myo, dice_avg, hd_epi, hd_myo, hd_avg]
        for j in range(max_column):
            cell = sheet.cell(max_row + 1, j + 1)
            cell.value = new_row[j]

        wb.save(excel_path_file)

def make_same_size(img_1, img_2):
        max_height = (max(img_1.shape[0], img_2.shape[0]))
        max_width = (max(img_1.shape[1], img_2.shape[1]))

        # the slices which adjusted resolution
        img_padding = np.zeros((max_height, max_width), dtype=np.uint8)
        if img_1.shape[0] == max_height and img_1.shape[1] == max_width:
            img_padding = img_1
        elif img_1.shape[0] < max_height and img_1.shape[1] < max_width:
            img_padding = zero_padding(np.asarray(img_1), constants.DIMENSION_HEIGHT, img_2.shape[0])
            img_padding = zero_padding(img_padding, constants.DIMENSION_WIDTH, img_2.shape[1])
        elif img_1.shape[0] < max_height:
            img_padding = zero_padding(np.asarray(img_1), constants.DIMENSION_HEIGHT, img_2.shape[0])
        elif img_1.shape[1] < max_width:
            img_padding = zero_padding(np.asarray(img_1), constants.DIMENSION_WIDTH, img_2.shape[1])

        res_img_1 = img_padding

        # reference slices which not adjusted resolution
        img_padding = np.zeros((max_height, max_width), dtype=np.uint8)
        if img_2.shape[0] == max_height and img_2.shape[1] == max_width:
            img_padding = img_2
        elif img_2.shape[0] < max_height and img_2.shape[1] < max_width:
            img_padding = zero_padding(np.asarray(img_2), constants.DIMENSION_HEIGHT, res_img_1.shape[0])
            img_padding = zero_padding(img_padding, constants.DIMENSION_WIDTH, res_img_1.shape[1])
        elif img_2.shape[0] < max_height:
            img_padding = zero_padding(np.asarray(img_2), constants.DIMENSION_HEIGHT, res_img_1.shape[0])
        elif img_2.shape[1] < max_width:
            img_padding = zero_padding(np.asarray(img_2), constants.DIMENSION_WIDTH, res_img_1.shape[1])
        res_img_2 = img_padding

        return res_img_1, res_img_2

def take_element(elem):
    return elem[1]

def take_list(lst):
    res = []
    for i in range(len(lst)):
        res.append(lst[i][0])
    return res