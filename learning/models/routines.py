import torch
from torchgeometry.losses.dice import DiceLoss
import warnings
import pandas as pd
import os
import openpyxl
import nibabel as nib
import numpy as np
import cv2
from learning.datasets import constants
from kornia.morphology import close
from medpy.metric import hd
from PIL import Image


def train_seg_model(dataloader, model, loss_fn, optimizer, device, mono_input=True):
    size = dataloader.dataset.__len__()
    if mono_input:
        for batch, (X, y) in enumerate(dataloader):
            X, y = X.to(device), y.to(device)
            # Compute prediction error
            pred = model(X)
            loss = loss_fn(pred, y)
            # Backpropagation
            optimizer.zero_grad()
            loss.backward()
            optimizer.step()
            loss, current, dice = loss.item(), batch * len(X), 1 - DiceLoss()(pred, y).item()  # nb : not true dice
            print(f"Training loss: {loss:>7f}, Dice: {dice:>0.2f} [{current:>5d}/{size:>5d}]")
    else:
        for batch, (X1, y1, X2, y2) in enumerate(dataloader):
            X1, y1 = X1.to(device), y1.to(device)
            X2, y2 = X2.to(device), y2.to(device)
            # Compute prediction error
            pred1, pred2 = model(X1, X2)
            loss1 = loss_fn(pred1, y1)
            loss2 = loss_fn(pred2, y2)
            loss = loss1 + loss2
            # Backpropagation
            optimizer.zero_grad()
            loss.backward()
            optimizer.step()
            loss, current, dice1, dice2 = loss.item(), batch * len(X1), 1 - DiceLoss()(pred1,
                                                                                       y1).item(), 1 - DiceLoss()(pred2,
                                                                                                                  y2).item()  # nb : not true dice
            print(
                f"Training loss: {loss:>7f}, Dice-Cine: {dice1:>0.2f}, Dice-DE: {dice2:>0.2f} [{current:>5d}/{size:>5d}]")
    return loss


from skimage import data, io, filters


def apply_seg_model(dataloader, model, loss_fn, device, save_im, save_scores, dataset, mono_input=True,
                    save_volume=False):
    size = dataloader.__len__()
    model.eval()
    num_class = dataloader.dataset.get_nbstruct()
    if mono_input:
        score_size = (num_class + 2)
    else:
        score_size = 2 * (num_class + 2)
    dice_scores = np.zeros(
        (size, score_size))  # automate later from nbstruct (ex : put function get_nbstruct in the dataloader)
    hausdorff_scores = np.zeros((size, score_size))

    test_loss, count = 0, 0
    aggregate_data = []
    with torch.no_grad():
        if mono_input:
            for X, y in dataloader:  # ! this version is sensitive to the batch size
                print(f"Case: [{count:>3d}/{size:>3d}]")
                X, y = X.to(device), y.to(device)
                pred = model(X)
                aggregate_data.append([(X, y, np.argmax(pred.long().numpy(),
                                                        axis=1))])  # aggregating all of the input data, GT , prediction

                header = dataloader.dataset.seg_nifti_headers[0]
                voxel_space = header['pixdim'][1:3]

                test_loss += loss_fn(pred, y).item()

                kernel = torch.ones(11, 11)  # post-processing
                pred = close(pred, kernel)
                pred = blob_detection(pred)

                dice_scores[count] = dice_score(pred, y, num_class)
                hausdorff_scores[count] = hausdorff(pred, y, num_class, dataloader.dataset.nifti_dim[0], voxel_space)
                count += 1
        else:
            for X1, y1, X2, y2 in dataloader:  # ! this version is sensitive to the batch size
                print(f"Case: [{count:>3d}/{size:>3d}]")
                X1, y1 = X1.to(device), y1.to(device)
                X2, y2 = X2.to(device), y2.to(device)
                pred1, pred2 = model(X1, X2)

                kernel = torch.ones(11, 11)  # post-processing
                pred1 = close(pred1, kernel)
                pred2 = close(pred2, kernel)

                test_loss += loss_fn(pred1, y1).item() + loss_fn(pred2, y2).item()

                pred1 = blob_detection(pred1)
                pred2 = blob_detection(pred2)
                aggregate_data.append([(X1, y1, np.argmax(pred1.long().numpy(), axis=1)),
                                       (X2, y2, np.argmax(pred2.long().numpy(),
                                                          axis=1))])  # aggregating all of the input data, GT , prediction for both modality

                dice_scores[count][:score_size // 2] = dice_score(pred1, y1, num_class)
                dice_scores[count][score_size // 2:] = dice_score(pred2, y2, num_class)

                # compute Hausdorff distance
                (header1, header2) = dataloader.dataset.seg_nifti_headers[0]

                voxel_space1 = header1['pixdim'][1:3]
                voxel_space2 = header2['pixdim'][1:3]
                hausdorff_scores[count][:score_size // 2] = hausdorff(pred1, y1, num_class,
                                                                      dataloader.dataset.nifti_dim[0], voxel_space1)
                hausdorff_scores[count][score_size // 2:] = hausdorff(pred2, y2, num_class,
                                                                      dataloader.dataset.nifti_dim[0], voxel_space2)
                count += 1
    test_loss /= size
    dice = np.nanmean(dice_scores, axis=0)
    hd = np.nanmean(hausdorff_scores, axis=0)

    if save_im:
        num_class = dataloader.dataset.get_nbstruct()
        if save_volume:
            dim = dataloader.dataset.nifti_dim
            if dataset == constants.DATASET_ACDC:
                save_ACDC_im(aggregate_data, dataloader.dataset.seg_nifti_headers, dim,
                             dataloader.dataset.patient_ids[0], num_class, device, len(dataloader))
            elif dataset == constants.DATASET_EMIDEC:
                save_EMIDEC_im(aggregate_data, dataloader.dataset.seg_nifti_headers, dim,
                               dataloader.dataset.patient_ids[0], num_class, device, len(dataloader))
            elif dataset == constants.DATASET_CINEDE:
                save_CINEDE_im(aggregate_data, dataloader.dataset.seg_nifti_headers, dim,
                               dataloader.dataset.patient_ids[0], num_class, device, len(dataloader))
            elif dataset == constants.DATASET_DE:
                save_DE_im(aggregate_data, dataloader.dataset.seg_nifti_headers, dim,
                           dataloader.dataset.patient_ids[0], num_class, device, len(dataloader))

    if save_scores:
        score_file_name = "results/scores.xlsx"
        if dataset == constants.DATASET_ACDC:
            new_row, df = save_ACDC_scores(dice, hd, test_loss)
        elif dataset == constants.DATASET_EMIDEC or dataset == constants.DATASET_DE:
            new_row, df = save_DE_scores(dice, hd, test_loss)
        elif dataset == constants.DATASET_CINEDE:
            new_row, df = save_CINDE_scores(dice, hd, test_loss)

        if not os.path.exists(score_file_name):
            df.to_excel(score_file_name, sheet_name='All', index=False)
        else:  # add new row
            wb = openpyxl.load_workbook(score_file_name)
            sheet = wb.active
            max_column = sheet.max_column
            max_row = sheet.max_row
            for j in range(max_column):
                cell = sheet.cell(max_row + 1, j + 1)
                cell.value = new_row[j]
            wb.save(score_file_name)
        print('Test Error: \n Dices: ', np.around(dice, 3), "\n")
        print('Hausdorff: ', np.around(hd, 3), f"Avg loss: {test_loss:>8f}" "\n")
    return test_loss


def save_ACDC_scores(dice, hd, test_loss):
    dice_rv = np.around(dice, 3)[1]
    dice_myo = np.around(dice, 3)[2]
    dice_lv = np.around(dice, 3)[3]
    mean_dice = np.around(dice, 3)[4]
    std_dice = np.around(dice, 3)[5]

    hd_rv = np.around(hd, 3)[1]  # hausdorff distance
    hd_myo = np.around(hd, 3)[2]
    hd_lv = np.around(hd, 3)[3]
    mean_hd = np.around(hd, 3)[4]
    std_hd = np.around(hd, 3)[5]

    df = pd.DataFrame({'dice_rv': [dice_rv],
                       'dice_myo': [dice_myo],
                       'dice_cv': [dice_lv],
                       'mean_dice': [mean_dice],
                       'std_dice': [std_dice],
                       'hd_rv': [hd_rv],
                       'hd_myo': [hd_myo],
                       'hd_cv': [hd_lv],
                       'mean_hd': [mean_hd],
                       'std_hd': [std_hd],
                       'Avg loss': [test_loss]
                       })
    new_row = [dice_rv, dice_myo, dice_lv, mean_dice, std_dice,
               hd_rv, hd_myo, hd_lv, mean_hd, std_hd,
               test_loss]
    return new_row, df


def save_DE_scores(dice, hd, test_loss):
    dice_cv = np.around(dice, 3)[1]
    dice_myo = np.around(dice, 3)[2]
    mean_dice = np.around(dice, 3)[3]
    std_dice = np.around(dice, 3)[4]
    hd_cv = np.around(hd, 3)[1]
    hd_myo = np.around(hd, 3)[2]
    mean_hd = np.around(hd, 3)[3]
    std_hd = np.around(hd, 3)[4]
    df = pd.DataFrame({'dice_cv': [dice_cv],
                       'dice_myo': [dice_myo],
                       'mean_dice': [mean_dice],
                       'std_dice': [std_dice],
                       'hd_cv': [hd_cv],
                       'hd_myo': [hd_myo],
                       'mean_hd': [mean_hd],
                       'std_hd': [std_hd],
                       'Avg loss': [test_loss]})
    new_row = [dice_cv, dice_myo, mean_dice, std_dice,
               hd_cv, hd_myo, mean_hd, std_hd,
               test_loss]
    return new_row, df


def save_CINDE_scores(dice, hd, test_loss):
    dice_cv_cine = np.around(dice, 3)[1]
    dice_myo_cine = np.around(dice, 3)[2]
    mean_dice_cine = np.around(dice, 3)[3]
    std_dice_cine = np.around(dice, 3)[4]
    dice_cv_de = np.around(dice, 3)[6]
    dice_myo_de = np.around(dice, 3)[7]
    mean_dice_de = np.around(dice, 3)[8]
    std_dice_de = np.around(dice, 3)[9]
    # hausdorff distance CINE
    hd_cv_cine = np.around(hd, 3)[1]
    hd_myo_cine = np.around(hd, 3)[2]
    mean_hd_cine = np.around(hd, 3)[3]
    std_hd_cine = np.around(hd, 3)[4]
    # hausdorff distance DE
    hd_cv_de = np.around(hd, 3)[6]
    hd_myo_de = np.around(hd, 3)[7]
    mean_hd_de = np.around(hd, 3)[8]
    std_hd_de = np.around(hd, 3)[9]

    df = pd.DataFrame({'dice_cv_cine': [dice_cv_cine],
                       'dice_myo_cine': [dice_myo_cine],
                       'mean_dice_cine': [mean_dice_cine],
                       'std_dice_cine': [std_dice_cine],
                       'dice_cv_de': [dice_cv_de],
                       'dice_myo_de': [dice_myo_de],
                       'mean_dice_de': [mean_dice_de],
                       'std_dice_de': [std_dice_de],
                       'hd_cv_cine': [hd_cv_cine],
                       'hd_myo_cine': [hd_myo_cine],
                       'mean_hd_cine': [mean_hd_cine],
                       'std_hd_cine': [std_hd_cine],
                       'hd_cv_de': [hd_cv_de],
                       'hd_myo_de': [hd_myo_de],
                       'mean_hd_de': [mean_hd_de],
                       'std_hd_de': [std_hd_de],
                       'Avg_Loss': [test_loss]})
    new_row = [dice_cv_cine, dice_myo_cine, mean_dice_cine, std_dice_cine,
               dice_cv_de, dice_myo_de, mean_dice_de, std_dice_de,
               hd_cv_cine, hd_myo_cine, mean_hd_cine, std_hd_cine,
               hd_cv_de, hd_myo_de, mean_hd_de, std_hd_de,
               test_loss]
    return new_row, df


def save_ACDC_im(aggregate_data, headers, nifti_dim, patient_ids, num_class, device, depth):
    path = "../../data/prediction_results"  # path to save prediction
    all_pred_slices = np.ones((256, 256, depth), dtype=np.uint8)
    all_gt_slices = np.ones((256, 256, depth), dtype=np.uint8)
    count = 0
    for data in aggregate_data:
        for indx, modality in enumerate(data):
            X = modality[0]
            y = modality[1]
            pred = modality[2]
            X, y = X.to(device), y.to(device)

            normalized_pred = cv2.normalize(pred[0].astype(np.uint8), np.zeros(pred[0].shape, dtype=np.uint8), 0,
                                            num_class - 1,
                                            cv2.NORM_MINMAX)
            normalized_y = cv2.normalize(y.long().numpy(), np.zeros(y[0].shape, dtype=np.uint8), 0, num_class - 1,
                                         cv2.NORM_MINMAX)

            all_pred_slices[:, :, count] = normalized_pred.astype(np.uint8)
            all_gt_slices[:, :, count] = normalized_y.astype(np.uint8)
        count += 1

        file_names = []
        for patient_id in patient_ids:
            file_names.append(patient_id)

        start_ed = 0
        for case in range(len(nifti_dim)):
            file_name = "prediction_" + str(file_names[case])
            volume_dim = (nifti_dim[case][0], nifti_dim[case][1], nifti_dim[case][2])

            subfolder = "ACDC"
            start_ed += 2 * nifti_dim[case - 1][2] if case > 0 else 0
            end_ed = start_ed + volume_dim[2]
            start_es = end_ed
            end_es = start_es + volume_dim[2]
            # decompose the data to ed and es (they were every other one, ed and es)
            ed = np.ones((256, 256, volume_dim[2]), dtype=np.uint8)
            es = np.ones((256, 256, volume_dim[2]), dtype=np.uint8)
            for slice_number in range(start_ed, end_es):
                if slice_number % 2 == 0:
                    ed[:, :, int((slice_number - start_ed) / 2)] = all_pred_slices[:, :, slice_number]
                else:
                    es[:, :, int((slice_number - start_ed - 1) / 2)] = all_pred_slices[:, :, slice_number]
            all_slices_ed_es = np.ones((256, 256, 2 * volume_dim[2]), dtype=np.uint8)
            all_slices_ed_es = np.concatenate((ed, es), axis=2)

            file_name_ed = file_name + "_ACDC_ed.nii.gz"
            save_nifti(all_slices_ed_es, headers[case], os.path.join(path, subfolder), 0, volume_dim[2], volume_dim,
                       file_name_ed)  # saving ed
            file_name_es = file_name + "_ACDC_es.nii.gz"
            save_nifti(all_slices_ed_es, headers[case], os.path.join(path, subfolder), volume_dim[2], 2 * volume_dim[2],
                       volume_dim,
                       file_name_es)  # saving es


def save_EMIDEC_im(aggregate_data, headers, nifti_dim, patient_ids, num_class, device, depth):
    path = "../../data/prediction_results"  # path to save prediction
    all_pred_slices = np.ones((256, 256, depth), dtype=np.uint8)
    all_gt_slices = np.ones((256, 256, depth), dtype=np.uint8)
    count = 0
    for data in aggregate_data:
        for indx, modality in enumerate(data):
            X = modality[0]
            y = modality[1]
            pred = modality[2]
            X, y = X.to(device), y.to(device)

            normalized_pred = cv2.normalize(pred[0].astype(np.uint8), np.zeros(pred[0].shape, dtype=np.uint8), 0,
                                            num_class - 1,
                                            cv2.NORM_MINMAX)
            normalized_y = cv2.normalize(y.long().numpy(), np.zeros(y[0].shape, dtype=np.uint8), 0, num_class - 1,
                                         cv2.NORM_MINMAX)

            all_pred_slices[:, :, count] = normalized_pred.astype(np.uint8)
            all_gt_slices[:, :, count] = normalized_y.astype(np.uint8)
        count += 1

        file_names = []
        for patient_id in patient_ids:
            file_names.append(patient_id)
        start = 0
        for case in range(len(nifti_dim)):
            file_name = "prediction_" + str(file_names[case])
            volume_dim = (nifti_dim[case][0], nifti_dim[case][1], nifti_dim[case][2])
            subfolder = "EMIDEC"
            file_suffix = "_EMIDEC.nii.gz"
            start += nifti_dim[case - 1][2] if case > 0 else 0
            end = start + volume_dim[2]

            file_name = file_name + file_suffix
            save_nifti(all_pred_slices, headers[case], os.path.join(path, subfolder), start, end, volume_dim, file_name)


def save_DE_im(aggregate_data, headers, nifti_dim, patient_ids, num_class, device, depth):
    path = "../../data/prediction_results"  # path to save prediction
    all_pred_slices = np.ones((256, 256, depth), dtype=np.uint8)
    all_gt_slices = np.ones((256, 256, depth), dtype=np.uint8)
    count = 0
    for data in aggregate_data:
        for indx, modality in enumerate(data):
            X = modality[0]
            y = modality[1]
            pred = modality[2]
            X, y = X.to(device), y.to(device)

            normalized_pred = cv2.normalize(pred[0].astype(np.uint8), np.zeros(pred[0].shape, dtype=np.uint8), 0,
                                            num_class - 1,
                                            cv2.NORM_MINMAX)
            normalized_y = cv2.normalize(y.long().numpy(), np.zeros(y[0].shape, dtype=np.uint8), 0, num_class - 1,
                                         cv2.NORM_MINMAX)

            all_pred_slices[:, :, count] = normalized_pred.astype(np.uint8)
            all_gt_slices[:, :, count] = normalized_y.astype(np.uint8)
        count += 1

        file_names = []
        for patient_id in patient_ids:
            file_names.append(patient_id)
        start = 0
        for case in range(len(nifti_dim)):
            file_name = "prediction_" + str(file_names[case])
            volume_dim = (nifti_dim[case][0], nifti_dim[case][1], nifti_dim[case][2])
            subfolder = "DE"
            file_suffix = "_DE.nii.gz"
            start += nifti_dim[case - 1][2] if case > 0 else 0
            end = start + volume_dim[2]

            file_name = file_name + file_suffix
            save_nifti(all_pred_slices, headers[case], os.path.join(path, subfolder), start, end, volume_dim, file_name)


def save_CINEDE_im(aggregate_data, headers, nifti_dim, patient_ids, num_class, device, depth):
    path = "../../data/prediction_results"  # path to save prediction
    all_pred_slices_cine = np.ones((256, 256, depth), dtype=np.uint8)
    all_gt_slices_cine = np.ones((256, 256, depth), dtype=np.uint8)

    all_pred_slices_de = np.ones((256, 256, depth), dtype=np.uint8)
    all_gt_slices_de = np.ones((256, 256, depth), dtype=np.uint8)
    count = 0
    for data in aggregate_data:
        for indx, modality in enumerate(data):
            X = modality[0]
            y = modality[1]
            pred = modality[2]
            X, y = X.to(device), y.to(device)

            if indx == 0:  # cine
                reference_pred_slices = all_pred_slices_cine
                reference_gt_slices = all_gt_slices_cine
            else:  # de
                reference_pred_slices = all_pred_slices_de
                reference_gt_slices = all_gt_slices_de

            normalized_pred = cv2.normalize(pred[0].astype(np.uint8), np.zeros(pred[0].shape, dtype=np.uint8), 0,
                                            num_class - 1,
                                            cv2.NORM_MINMAX)
            normalized_y = cv2.normalize(y.long().numpy(), np.zeros(y[0].shape, dtype=np.uint8), 0, num_class - 1,
                                         cv2.NORM_MINMAX)

            reference_pred_slices[:, :, count] = normalized_pred.astype(np.uint8)
            reference_gt_slices[:, :, count] = normalized_y.astype(np.uint8)
        count += 1

        file_names = []
        for patient_id in patient_ids:
            file_names.append(patient_id)
        start = 0
        for case in range(len(nifti_dim)):
            file_name = "prediction_" + str(file_names[case])
            volume_dim = (nifti_dim[case][0], nifti_dim[case][1], nifti_dim[case][2])
            subfolder = "CINEDE"
            start += nifti_dim[case - 1][2] if case > 0 else 0
            end = start + volume_dim[2]
            file_suffix = "_CINE.nii.gz"
            full_file_name = file_name + file_suffix
            save_nifti(all_pred_slices_cine, headers[case][0], os.path.join(path, subfolder), start, end, volume_dim,
                       full_file_name)  # saving Cine prediction

            file_suffix = "_DE.nii.gz"
            full_file_name = file_name + file_suffix
            save_nifti(all_pred_slices_de, headers[case][1], os.path.join(path, subfolder), start, end, volume_dim,
                       full_file_name)  # saving Cine prediction


def save_nifti(all_slices, header, path, start, end, volume_dim, file_name):
    if not os.path.exists(path):
        os.makedirs(path)
    case_slices = np.zeros(volume_dim, dtype=np.uint8)
    for slice_number in range(start, end):
        case_slices[:, :, slice_number - start] = resize(all_slices[:, :, slice_number], volume_dim, 0)

    if (header['dim'][1], header['dim'][2], header['dim'][3]) != volume_dim:  # to match with ITK-Snap
        case_slices = case_slices.transpose(1, 0, 2)

    pred_nifti = nib.Nifti1Image(case_slices, np.eye(4), header=header.copy())
    nib.save(pred_nifti, os.path.join(path, file_name))


def resize(image, shape, interp_option):
    image_resized = np.array(
        Image.fromarray(image.astype(np.uint8)).resize((shape[1], shape[0]), resample=interp_option))
    return image_resized


def dice_score(prediction, gt, num_class):
    prediction = prediction.long().numpy()
    gt = gt.long().numpy()
    nbc = gt.shape[0]
    mask = np.argmax(prediction, axis=1)
    nbdim = num_class  # to automate from the dataloader
    dices = np.zeros((nbc, nbdim + 2))
    for c in range(nbc):
        for s in range(nbdim):
            if np.sum(gt[c] == s) == 0:
                dices[c, s] = np.NaN  # structure not present in the gt
            else:
                dices[c, s] = np.sum((mask[c] == s) * (gt[c] == s) * 2.0) / (np.sum(mask[c] == s) + np.sum(gt[c] == s))

        dices[c, -2] = np.mean(dices[c, 1:-2])
        dices[c, -1] = np.std(dices[c, 1:-2])

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", category=RuntimeWarning)  # ignore case with all nans
        dices = np.nanmean(dices, axis=0)
    return dices


def hausdorff(prediction, gt, num_class, shape, voxel_space):
    prediction = prediction.cpu().long().numpy()
    gt = gt.cpu().long().numpy()
    nbc = gt.shape[0]
    mask = np.argmax(prediction, axis=1)
    nbdim = num_class
    hausdorffs = np.zeros((nbc, nbdim + 2))
    for c in range(nbc):
        for s in range(nbdim):
            gt_copy = gt[c].copy()
            mask_copy = mask[c].copy()

            if np.sum(gt[c] == s) == 0:
                hausdorffs[c, s] = np.NaN  # structure not present in the gt
            else:
                gt_copy[gt_copy != s] = 0
                gt_copy[gt_copy == s] = 1

                mask_copy[mask_copy != s] = 0
                mask_copy[mask_copy == s] = 1

                resized_gt = resize(gt_copy, shape, 0)
                resized_mask = resize(mask_copy, shape, 0)

                if np.count_nonzero(resized_mask) == 0:  # structure was not predicted
                    hausdorffs[c, s] = np.NaN
                else:
                    hausdorffs[c, s] = hd(resized_mask, resized_gt, voxel_space)

        hausdorffs[c, -2] = np.mean(hausdorffs[c, 1:-2])
        hausdorffs[c, -1] = np.std(hausdorffs[c, 1:-2])

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", category=RuntimeWarning)  # ignore case with all nans
        hausdorffs = np.nanmean(hausdorffs, axis=0)
    return hausdorffs


def blob_detection(pred):
    params = cv2.SimpleBlobDetector_Params()
    # Filter by Circularity
    params.filterByCircularity = True
    params.minCircularity = 0.1

    detector = cv2.SimpleBlobDetector_create(params)

    pred_np = np.argmax(pred, axis=1)[0].long().numpy().astype(np.uint8)
    pred_np[pred_np == 1] = 128
    pred_np[pred_np == 2] = 255
    keypoints = detector.detect(pred_np)

    result = torch.Tensor(1, 3, 256, 256)
    result[:, :, :, :] = pred[0, :, 10:11, 10:11]  # default value for pixels

    if len(keypoints) > 0:
        for kp in keypoints:
            r = int(0.8 * kp.size)
            (x, y) = np.int0(kp.pt)
            result[:, :, y - r:y + r, x - r:x + r] = pred[:, :, y - r:y + r, x - r:x + r]
    return result
